from openpyxl import load_workbook
from openpyxl import Workbook

class ExcelDetailsBase():
    # wb
    # det_sheet
    # sb_sheet
    # col_det_numbers
    # col_sb_numbers
    
    def load(self, FILENAME):
        self.wb = Workbook()
        self.det_sheet = self.wb.active
        
        # Загрузка рабочей книги
        # self.workbook = load_workbook(filename=FILENAME, read_only=False)
        
        # Получение листов
        # self.det_sheet = self.workbook[1]
        # self.sb_sheet = self.workbook[2]
        
        # Получение номера последней строки
        # det_last_row = self.det_sheet.max_row
        # sb_last_row = self.sb_sheet.max_row
        
        # Чтение столбцов из книги в списки
        # self.col_det_numbers = self.det_sheet['B1:B'+str(det_last_row)]
        # self.col_sb_numbers = self.sb_sheet['B1:B'+str(sb_last_row)]
        
    # def find_row_by_number(number):
        # for i in range(0, )
    
    def close(self):
        # закрываем книгу после прочтения
        self.wb.save('test.xlsx')
        self.wb.close()
    
    def write_info(self, info):
        #self.det_sheet.append(info)
        row_number = 2
        for file_info in info:
            for key in file_info:
                if key != 'Filepath': 
                    self.det_sheet.cell(row=row_number, column=1).value = file_info['Filepath']
                    self.det_sheet.cell(row=row_number, column=2).value = key
                    self.det_sheet.cell(row=row_number, column=3).value = file_info[key]
                    row_number += 1
        








