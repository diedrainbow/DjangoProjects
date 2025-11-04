from .models import Detail, SB, Material
from datetime import date, datetime
import time
from django.utils import timezone

import os
import codecs
import csv
from openpyxl import load_workbook


class TimeAnalis():
    prev_t = time.time()
    times = dict()
    
    def time_update(self, key):
        if key not in self.times: self.times.update( {key:0.0} )
        t = time.time()
        self.times[key] += t - self.prev_t
        self.prev_t = t
        
    def get_string(self):
        s = '=== Times ===\n'
        for key in self.times:
            s += key + ': ' + str(self.times[key]) + ';\n'
        s += '============='
        return s
        
    def get_sum(self):
        s = 0.0
        for key in self.times:
            s += self.times[key]
        return str(s)



def load_SB_from_xlsx(FILENAME):
    # Загрузка рабочей книги
    wb = load_workbook(filename=FILENAME, read_only=True)
    # Получение активного листа
    sheet = wb['Список сборочных']
    
    # Получение номера последней строки
    last_row_number = sheet.max_row
    print(f"Номер последней строки с данными: {last_row_number}")
    
    for row_number in range(4, last_row_number):
        sb_number = sheet.cell(row=row_number, column=2).value
        if sb_number == "": continue
        
        try:
            sb = SB.objects.get(number = sb_number)
        except SB.DoesNotExist:
            sb = SB()
            
        sb.number       = sb_number
        sb.name         = sheet.cell(row=row_number, column=3).value
        
        composition = sheet.cell(row=row_number, column=4).value
        if composition == None:
            composition = ""
        else:
            sb.composition  = composition.replace("_x000D_", "; ")
        
        sb.actual_date  = DateFromOpenpyxl(sheet.cell(row=row_number, column=5).value)
        sb.comment      = ""
        
        cdw_file_name = sheet.cell(row=row_number, column=6).value
        if cdw_file_name == "" or cdw_file_name == None or sheet.cell(row=row_number, column=8).value == "Нет файла":
            sb.cdw_file_name = ""
            sb.cdw_file_folder = ""
            sb.cdw_file_date = None
            sb.cdw_valid = 0.0
        else:
            sb.cdw_file_name = cdw_file_name
            sb.cdw_file_folder = sheet.cell(row=row_number, column=7).value
            sb.cdw_file_date = DateFromOpenpyxl(sheet.cell(row=row_number, column=8).value)
            
            if sheet.cell(row=row_number, column=6).fill is None:
                color_index = "00000000"
            else:
                color_index = sheet.cell(row=row_number, column=6).fill.start_color.index
            
            if color_index == "00000000":
                sb.cdw_valid = 1.0
            else:
                sb.cdw_valid = 0.5
        
        sb.save()
        print(f"Прогресс: {row_number} строк", end='\r')


def load_details_from_xlsx(FILENAME):
    # Загрузка рабочей книги
    wb = load_workbook(filename=FILENAME, read_only=True)
    # Получение активного листа
    sheet = wb['Список деталей']
    
    # Получение номера последней строки
    last_row_number = sheet.max_row
    print(f"Номер последней строки с данными: {last_row_number}")
    
    # Чтение столбцов из книги в списки
    last_row_number_str = str(last_row_number)
    col_material = sheet['A1:A'+last_row_number_str]
    print("col_material load")
    col_number = sheet['B1:B'+last_row_number_str]
    print("col_number load")
    col_name = sheet['C1:C'+last_row_number_str]
    print("col_name load")
    col_grave = sheet['D1:D'+last_row_number_str]
    print("col_grave load")
    col_size = sheet['E1:E'+last_row_number_str]
    print("col_size load")
    col_marsh = sheet['F1:F'+last_row_number_str]
    print("col_marsh load")
    col_poddon = sheet['G1:G'+last_row_number_str]
    print("col_poddon load")
    col_comment = sheet['H1:H'+last_row_number_str]
    print("col_comment load")
    col_date = sheet['I1:I'+last_row_number_str]
    print("col_date load")
    
    col_frw_name = sheet['J1:J'+last_row_number_str]
    print("col_frw_name load")
    col_frw_dir = sheet['K1:K'+last_row_number_str]
    print("col_frw_dir load")
    col_frw_date = sheet['L1:L'+last_row_number_str]
    print("col_frw_date load")
    col_frw_parts = sheet['P1:P'+last_row_number_str]
    print("col_frw_parts load")
    
    col_cdw_name = sheet['M1:M'+last_row_number_str]
    print("col_cdw_name load")
    col_cdw_dir = sheet['N1:N'+last_row_number_str]
    print("col_cdw_dir load")
    col_cdw_date = sheet['O1:O'+last_row_number_str]
    print("col_cdw_date load")
    
    col_stage1 = sheet['Q1:Q'+last_row_number_str]
    print("col_stage1 load")
    col_descr1 = sheet['R1:R'+last_row_number_str]
    print("col_descr1 load")
    col_stage2 = sheet['S1:S'+last_row_number_str]
    print("col_stage2 load")
    col_descr2 = sheet['T1:T'+last_row_number_str]
    print("col_descr2 load")
    col_stage3 = sheet['U1:U'+last_row_number_str]
    print("col_stage3 load")
    col_descr3 = sheet['V1:V'+last_row_number_str]
    print("col_descr3 load")
    col_stage4 = sheet['W1:W'+last_row_number_str]
    print("col_stage4 load")
    col_descr4 = sheet['X1:X'+last_row_number_str]
    print("col_descr4 load")
    
    # Лог-файл 
    #logfile = open(r"d:\DjangoProject\DjangoProdactionControl\logfiles\load_details_from_xlsx(" + str(time.time()) + ").txt", "w")
    logfile = open(r"d:\django\DjangoProdactionControl\logfiles\load_details_from_xlsx(" + str(time.time()) + ").txt", "w", encoding="utf-8")
    logfile.write(FILENAME)
 
    # Объект анализа времени выполнения отдельных сегментов кода
    ta = TimeAnalis()
    ta.time_update('На подготовку массивов')
    
    for row_number in range(3, last_row_number-1): # 
    
        try:
    
            detail_number = str(col_number[row_number][0].value)
            if detail_number == "" or detail_number == 'NULL' : continue
            detail = Detail()
            
            ta.time_update('На печать инфы и получение номера детали')
            
            detail.material     = GetMaterialFromString(col_material[row_number][0].value)
            
            ta.time_update('На поиск или создание материала')
            
            detail.number       = detail_number
            detail.name         = str(col_name[row_number][0].value)
            detail.grave        = tostr(col_grave[row_number][0].value)
            detail.size         = str(col_size[row_number][0].value)
            detail.marshrut     = str(col_marsh[row_number][0].value)
            detail.poddon       = str(col_poddon[row_number][0].value)
            detail.comment      = tostr(col_comment[row_number][0].value)
            
            ta.time_update('На чтение нескольких ячеек')
            
            detail.actual_date  = DateFromOpenpyxl(col_date[row_number][0].value)
            
            ta.time_update('На чтение и преобразование даты')
            
            frw_file_name = col_frw_name[row_number][0].value
            if frw_file_name == "" or frw_file_name == None or col_frw_date[row_number][0].value == "Нет файла":
                detail.frw_file_name = ""
                detail.frw_file_folder = ""
                detail.frw_file_date = None
                detail.frw_valid = 0.0
            else:
                detail.frw_file_name = frw_file_name
                detail.frw_file_folder = col_frw_dir[row_number][0].value
                detail.frw_file_date = DateFromOpenpyxl(col_frw_date[row_number][0].value)
                
                #fill = sheet.cell(row=row_number+1, column=10).fill
                fill = col_frw_name[row_number][0].fill
                if fill is None: color_index = "00000000"
                else: color_index = fill.start_color.index
                    
                if color_index == "00000000": detail.frw_valid = 1.0
                else: detail.frw_valid = 0.5
            
            ta.time_update('На чтение и операции с ФРВ')
            
            cdw_file_name = col_cdw_name[row_number][0].value
            if cdw_file_name == "" or cdw_file_name == None or col_cdw_date[row_number][0].value == "Нет файла":
                detail.cdw_file_name = ""
                detail.cdw_file_folder = ""
                detail.cdw_file_date = None
                detail.cdw_valid = 0.0
            else:
                detail.cdw_file_name = cdw_file_name
                detail.cdw_file_folder = col_cdw_dir[row_number][0].value
                detail.cdw_file_date = DateFromOpenpyxl(col_cdw_date[row_number][0].value)
                
                #fill = sheet.cell(row=row_number+1, column=13).fill
                fill = col_cdw_name[row_number][0].fill
                if fill is None: color_index = "00000000"
                else: color_index = fill.start_color.index
                    
                if color_index == "00000000": detail.cdw_valid = 1.0
                else: detail.cdw_valid = 0.5
             
            ta.time_update('На чтение и операции с СДВ')
            
            detail.parts = col_frw_parts[row_number][0].value
            
            detail.stages = tostr(col_stage1[row_number][0].value) + ';'
            detail.stages += tostr(col_stage2[row_number][0].value) + ';'
            detail.stages += tostr(col_stage3[row_number][0].value) + ';'
            detail.stages += tostr(col_stage4[row_number][0].value) + ';'
            detail.stages += ';;'
            
            detail.descriptions = tostr(col_descr1[row_number][0].value) + ';'
            detail.descriptions += tostr(col_descr2[row_number][0].value) + ';'
            detail.descriptions += tostr(col_descr3[row_number][0].value) + ';'
            detail.descriptions += tostr(col_descr4[row_number][0].value) + ';'
            detail.descriptions += ';;'
            
            ta.time_update('На чтение Стадий и Расшифровок')
            
            try:
                new_detail = Detail.objects.get(number = detail_number)
            except Detail.DoesNotExist:
                new_detail = Detail()
            
            log_str = new_detail.all_propertys_overrade_and_save(detail)
            
            if log_str != '':
                log_str = '\n id: ' + str(new_detail.id) + log_str
                logfile.write(safe_tostr(log_str))
            
            ta.time_update('На поиск и сохранение детали в базе')
            
            print(f"Прогресс: {row_number} строк. Время: " + ta.get_sum(), end='\r')
        
        except Exception as ex:
            log_str = '\n === ERROR === row: ' + str(row_number) + '; ' + str(ex)
            logfile.write(log_str)
            print(log_str)
        
    # закрываем книгу после прочтения
    wb.close()
    print('\n', ta.get_string(), '\n')
    
    logfile.write('\n')
    logfile.write(ta.get_string())
    logfile.close()

def tostr(val):
    if val == None: return ''
    if isinstance(val, str): return val
    return str(val)

def safe_tostr(val):
    try:
        # Словарь замен для распространенных проблемных символов
        char_replacements = {
            '\xbd': '½',  # Vulgar fraction one half
            '\xbc': '¼',  # Vulgar fraction one quarter
            '\xbe': '¾',  # Vulgar fraction three quarters
            '\xb0': '°',  # Degree symbol
        }
        
        result = tostr(val) + ';'
        for bad_char, good_char in char_replacements.items():
            result = result.replace(bad_char, good_char)
            
        return result
    except UnicodeDecodeError:
        print("=========================", val)
        # Пробуем разные кодировки
        for encoding in ['utf-8', 'latin-1', 'cp1251', 'cp1252']:
            try:
                if isinstance(val, bytes):
                    return val.decode(encoding) + ';'
                else:
                    return str(val).encode('utf-8').decode(encoding) + ';'
            except:
                continue
            
            # Если ничего не помогло, заменяем проблемные символы
            return str(val).encode('utf-8', errors='replace').decode('utf-8') + ';'
    

def load_from_csv(FILENAME):
    try:
        with open(FILENAME, "r", newline="") as file:
            reader = csv.reader(file, delimiter=';')
            #header = next(reader)
            #print(header)
            count = 0
            
            for row in reader:
                try:
                    detail = Detail.objects.get(number = row[1])
                except Detail.DoesNotExist:
                    detail = Detail()
                
                # Общие данные о детали
                detail.material_name    = GetMaterialNameFromString(row[0])
                detail.number           = row[1]
                detail.name             = row[2]
                detail.grave            = row[3]
                detail.size             = row[4]
                detail.marshrut         = row[5]
                detail.poddon           = row[6]
                detail.comment          = row[7]
                act_date = DateTimeFromString(row[8])
                if act_date == None: act_date = datetime(2000, 1, 1, 1, 1, 1)
                detail.actual_date      = act_date
                
                # Ссылки на файлы чертежа и фрагмента
                if row[9] == "":
                    detail.frw_file_name    = ""
                    detail.frw_file_folder  = ""
                    detail.frw_file_date    = None
                    detail.frw_file_parts   = ""
                    detail.frw_valid        = 0.0
                else:    
                    detail.frw_file_name    = row[9].replace("[,]", ";")
                    detail.frw_file_folder  = row[10].replace("[,]", ";")
                    detail.frw_file_date    = DateTimeFromString(row[11])
                    detail.frw_file_parts   = row[15].replace("[,]", ";")
                    detail.frw_valid        = 1.0
                
                if row[12] == "":
                    detail.cdw_file_name    = ""
                    detail.cdw_file_folder  = ""
                    detail.cdw_file_date    = None
                    detail.cdw_valid        = 0.0
                else:
                    detail.cdw_file_name    = row[12].replace("[,]", ";")
                    detail.cdw_file_folder  = row[13].replace("[,]", ";")
                    detail.cdw_file_date    = DateTimeFromString(row[14])
                    detail.cdw_valid        = 1.0
                    
                # Технологические операции
                detail.stages           = row[16]+";"+row[18]+";"+row[20]+";"+row[22]+";"+row[24]+";"+";"
                detail.times            = ";;;;;;"
                detail.descriptions     = row[17]+";"+row[19]+";"+row[21]+";"+row[23]+";"+row[25]+";"+";"
                
                detail.save()
                count += 1
                print(f"Прогресс: {count} строк", end='\r')
            
    except Exception as ex:
        print("============================== ERROR ===============================")
        print(count, " строка => ", ex)
        print("====================================================================")
        return 
    
    print("\n === Complit")
    


def DateFromOpenpyxl(date):
    if date == None:
        date = datetime(2000, 1, 1, 1, 1, 1)
    return timezone.make_aware(date)
    

def DateTimeFromString(date_string):
    if date_string == "": return None
    # Преобразование строки в объект datetime
    try:
        date_object = datetime.strptime(date_string, "%d.%m.%Y")
    except ValueError as e:
        try:
            date_object = datetime.strptime(date_string, "%d.%m.%Y %H:%M")
        except ValueError as e:
            date_object = datetime(2000, 1, 1, 1, 1, 1)
    return timezone.make_aware(date_object)


def GetMaterialFromString(material_string):
    if material_string == '' or material_string == None : return None
    
    material, created = Material.objects.get_or_create(original = material_string)
    if created:
        material.SetMaterialFromString( material_string )
        print("new material "+material.original)
        material.save()
    
    # try:
        # material = Material.objects.get(original = material_string)
    # except Material.DoesNotExist:
        # material = Material()
        # material.SetMaterialFromString( material_string )
        # print("new material "+material.original)
        # material.save()
    
    return material



